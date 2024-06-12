from tkinter import filedialog
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill
from itertools import chain

global EVENT_COLOR
EVENT_COLOR = "F0FC5A"
global disabled_sections
disabled_sections = []
global HIGH_AMPLITUDE_MINIMUM
HIGH_AMPLITUDE_MINIMUM_VALUE = 75
global HIGH_AMPLITUDE_LENGTH
HIGH_AMPLITUDE_MINIMUM_PATTERN_LENGTH = 4
global HIGH_AMPLITUDE_MINIMUM_LENGTH_CM
HIGH_AMPLITUDE_MINIMUM_LENGTH_MM = 100


def custom_sort(item):
    order = ["Ascending", "Transverse", "Descending", "Sigmoid", "Rectum"]
    return order.index(item)

def remove_disabled_sections(section):
    global disabled_sections
    if section in disabled_sections:
        disabled_sections.remove(section)
    disabled_sections = sorted(disabled_sections, key=custom_sort)

def add_disabled_sections(section):
    global disabled_sections
    if section not in disabled_sections:
        disabled_sections.append(section)
    disabled_sections = sorted(disabled_sections, key=custom_sort)

def reset_disabled_sections():
    global disabled_sections
    disabled_sections = []

def exportToXlsx(data, file_name, sliders, events, settings_sliders, first_event_text):
    # Split the file path into the base name and extension
    base_name, ext = file_name.rsplit('.', 1)

    # Create the new file path with '_analysis' appended to the base name
    new_file_name = f"{base_name}_analysis.xlsx"
    
    # Write the DataFrame to an Excel file
    try:
        data.to_excel(new_file_name, index=False)
        insertEmptyRows(new_file_name, 9)
        mergeAndColorCells(new_file_name, sliders)
        event_names = []
        for time, event_name in events.items():
            event_names.append(event_name)
            total_seconds = time // 10  # Convert deciseconds to seconds
            hour, remainder = divmod(total_seconds, 3600)  # Calculate hours
            minute, second = divmod(remainder, 60)  # Calculate minutes and seconds
            addEventNameAtGivenTime(new_file_name, hour, minute, second, event_name)
        wb = assignSectionsBasedOnStartSection(new_file_name, sliders, event_names, settings_sliders, first_event_text.get())
        file_name = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")], initialfile=new_file_name)
        wb.save(file_name)
        print(f"Data successfully exported to {new_file_name}")
    except Exception as e:
        print(f"Error exporting data to Excel: {e}")

def getSliderValues(sliders):
    list_with_slider_tuples = []
    for i in range(len(sliders)):
        value1 = -1
        value2 = -1
        for element in sliders[i].get():
            if value1 == -1:
                value1 = round(element)
            else:
                value2 = round(element)
        slider_tuple = (value1,value2)
        list_with_slider_tuples.append(slider_tuple)
    return list_with_slider_tuples

def mergeAndColorCells(file_name, sliders):
    # Load the workbook and select the active sheet
    wb = load_workbook(file_name)
    ws = wb.active
    
    # Define the sections and their corresponding colors
    sections = ["Ascending", "Transverse", "Descending", "Sigmoid", "Rectum"]
    sliders = getSliderValues(sliders)

    for section in disabled_sections:
        sections.remove(section)
        sliders.pop(0)

    colors = {
        "Ascending": "A9D08E",
        "Transverse": "BDD7EE",
        "Descending": "F8CBAD",
        "Sigmoid": "D9D9D9",
        "Rectum": "B1A0C7"
    }

    # Merge cells and color them for each section
    for i, (start, end) in enumerate(sliders):
        start_col = get_column_letter(start + 11)
        end_col = get_column_letter(end + 11)
        ws.merge_cells(f'{start_col}22:{end_col}22')
        cell = ws[f'{start_col}22']
        cell.value = sections[i]
        cell.alignment = Alignment(horizontal='center', vertical='center')
        fill = PatternFill(start_color=colors[sections[i]], end_color=colors[sections[i]], fill_type="solid")
        cell.fill = fill
        for col in range(start + 11, end + 12):
            ws[f'{get_column_letter(col)}22'].fill = fill

    # Save the workbook
    wb.save(file_name)

def addEventNameAtGivenTime(file_name, hour, minute, second, event_name):
    # Load the workbook and select the active sheet
    wb = load_workbook(file_name)
    ws = wb.active
    
    # Find the insertion row based on the specified hour, minute, and second
    insertion_row = None
    for row in range(24, ws.max_row + 1):
        cell_hour = ws.cell(row=row, column=2).value
        cell_minute = ws.cell(row=row, column=3).value
        cell_second = ws.cell(row=row, column=4).value
        
        if (int(cell_hour) > int(hour)) or (int(cell_hour) == int(hour) and int(cell_minute) > int(minute)) or (int(cell_hour) == int(hour) and int(cell_minute) == int(minute) and int(cell_second) >= int(second)):
            insertion_row = row
            break
    
    if insertion_row is None:
        insertion_row = ws.max_row + 1
    
    # Insert a new row at the identified position
    ws.insert_rows(insertion_row)
    
    for col in range(1, 11):
        ws.cell(row=insertion_row, column=col, value=event_name)
        ws.cell(row=insertion_row, column=col).fill = PatternFill(start_color=EVENT_COLOR, end_color=EVENT_COLOR, fill_type="solid")
    # Fill in the event details
    # ws.cell(row=insertion_row, column=1, value=insertion_row - 1)
    ws.cell(row=insertion_row, column=2, value=hour)
    ws.cell(row=insertion_row, column=3, value=minute)
    ws.cell(row=insertion_row, column=4, value=second)

    # Save the workbook
    wb.save(file_name)

def insertEmptyRows(file_name, amount):
    wb = load_workbook(file_name)
    ws = wb.active

    for i in range (amount):
        ws.insert_rows(13 + i)
    
    wb.save(file_name)

def assignSectionsBasedOnStartSection(file_name, sliders, event_names, settings_sliders, first_event_text):
    # Load the workbook and select the active sheet
    wb = load_workbook(file_name)
    ws = wb.active

    # Define the sections and their corresponding colors
    sections = ["Ascending", "Transverse", "Descending", "Sigmoid", "Rectum"]
    colors = {
        "Ascending": "A9D08E",
        "Transverse": "BDD7EE",
        "Descending": "F8CBAD",
        "Sigmoid": "D9D9D9",
        "Rectum": "B1A0C7",
        "Ascending tot in Rectum": "81BA5A",
        "Transverse tot in Rectum": "81B2DF",
        "Descending tot in Rectum": "F2A16A",
        "Sigmoid tot in Rectum": "BEBEBE"
    }

    # Get the slider values
    sliders = getSliderValues(sliders)
    distance_between_sensors = int(round(settings_sliders[0].get()))

    counters = {}
    length_counters = {}
    high_amplitude_counters = {}
    
    if first_event_text.strip(' ') == '':
        first_event_text = "Post-Wake"

    counters.update({str(first_event_text): {}})
    length_counters.update({str(first_event_text): {}})
    high_amplitude_counters.update({str(first_event_text): {}})
    for event_name in event_names:
        counters.update({event_name: {}})
        length_counters.update({event_name: {}})
        high_amplitude_counters.update({event_name: {}})

    counter_template = {
        "Ascending": 0,
        "Transverse": 0,
        "Descending": 0,
        "Sigmoid": 0,
        "Rectum": 0,
        "Ascending tot in Rectum": 0,
        "Transverse tot in Rectum": 0,
        "Descending tot in Rectum": 0,
        "Sigmoid tot in Rectum": 0,
    }

    length_counter_template = {
        "aantal long s": 0,
        "aantal short s": 0,
        "aantal long r": 0,
        "aantal short r": 0,
        "aantal long a": 0,
        "aantal short a": 0,
    }

    high_amplitude_counters_template = {
        "HAPCs": 0,
        "HARPCs": 0,
    }

    length_counter = length_counter_template.copy()

    keys_to_remove = []
    for section in disabled_sections:
        sections.remove(section)
        sliders.pop(0)
        for key in counter_template.keys():
            if section in key:
                keys_to_remove.append(key)

    for key in keys_to_remove:
        del counter_template[key]
    counter = counter_template.copy()

    high_amplitude_counter = high_amplitude_counters_template.copy()
    
    # Iterate over each row starting from row 24
    for row in range(24, ws.max_row + 1):
        pattern = ws.cell(row=row, column=6).value
        length = ws.cell(row=row, column=10).value
        if(isinstance(length, int)):
            pattern_type = "long" if distance_between_sensors * length > 100 else "short"
            length_counter[f"aantal {pattern_type} {pattern}"] += 1
            HIGH_AMPLITUDE_DISTANCE = distance_between_sensors * length
        amount_of_high_amp_in_row = 0
        skip = False
        for col in range(12, ws.max_column + 1):  # Adjust the starting column index if needed
            cell_value = ws.cell(row=row, column=col).value

            if isinstance(cell_value, (int, float)) and cell_value > 0 and not skip:
                # Determine the section based on the column index and slider ranges
                for i, (start, end) in enumerate(sliders[:-1]):
                    section = sections[i]
                    rectum_start = sliders[4-len(disabled_sections)][0]
                    if(ws.cell(row=row, column=col).value != None and start + 11 <= col <= end + 11):
                        # Check if everything from the section to the rectum is filled in
                        if all(ws.cell(row=row, column=col2).value != None for col2 in range(col,  rectum_start + 12)):
                            counter.update({f'{section} tot in Rectum': counter[f'{section} tot in Rectum'] + 1})
                            break
                                            
                for i, (start, end) in enumerate(sliders):
                    if start + 11 <= col <= end + 11:
                        section = sections[i]
                        counter.update({section: counter[section] + 1})
                        length_cell = ws.cell(row=row, column=10)  #column 10 has the length data
                        fill = PatternFill(start_color=colors[section], end_color=colors[section], fill_type="solid")
                        length_cell.fill = fill
                        break   
                skip = True
            if(isinstance(cell_value, (int, float)) and cell_value > HIGH_AMPLITUDE_MINIMUM_VALUE):
                amount_of_high_amp_in_row += 1

        if amount_of_high_amp_in_row >= HIGH_AMPLITUDE_MINIMUM_PATTERN_LENGTH and HIGH_AMPLITUDE_DISTANCE >= HIGH_AMPLITUDE_MINIMUM_LENGTH_MM:
            pattern_direction = ws.cell(row=row, column=6).value
            if pattern_direction == "r":
                high_amplitude_counter.update({"HARPCs": high_amplitude_counter["HARPCs"] + 1})
            elif pattern_direction == "a":
                high_amplitude_counter.update({"HAPCs": high_amplitude_counter["HAPCs"] + 1})
            elif pattern_direction == "s":
                print("Wat moet met sync gebeuren?")
        #check if the cell value is a string instead of an int, if so add counter to counters (event name as key and counter as value) and reset each value to 0
        next_row = ws.cell(row=row + 1, column=1).value
        if isinstance(ws.cell(row=row, column=1).value, str) or next_row == None:                
            #subtract the "section" with "section tot in Rectum" to get the correct amount of in sections
            for section in sections[:-1]:
                counter[section] -= counter[f'{section} tot in Rectum']

            # add the counter to the first empty value from a key in the counters dictionary
            for event in counters.keys():
                if counters[event] == {}:
                    counters[event] = counter
                    break
            for event in length_counters.keys():
                if length_counters[event] == {}:
                    length_counters[event] = length_counter
                    break
            for event in high_amplitude_counters.keys():
                if high_amplitude_counters[event] == {}:
                    high_amplitude_counters[event] = high_amplitude_counter
                    break
            counter = counter_template.copy()
            length_counter = length_counter_template.copy()
            high_amplitude_counter = high_amplitude_counters_template.copy()

    #Section names added to the left of table
    row = 3
    for section in counter.keys():
        ws.cell(row=row, column=19, value=section)
        fill = PatternFill(start_color=colors[section], end_color=colors[section], fill_type="solid")
        ws.cell(row=row, column=19).fill = fill
        row += 1
    row += 1
    #Pattern types (eg. short s) added to the left of table
    length_counter_row_start = row
    for pattern in list(chain(length_counter.keys(), high_amplitude_counter.keys())):
        ws.cell(row=row, column=19, value=pattern)
        fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
        ws.cell(row=row, column=19).fill = fill
        row += 1
    
    #Fill in the counters
    column = 20
    for (event,value) in counters.items():
        row = 2
        ws.cell(row=row, column=column, value=event)
        fill = PatternFill(start_color=EVENT_COLOR, end_color=EVENT_COLOR, fill_type="solid")
        ws.cell(row=row, column=column).fill = fill
        row += 1
        for (section,value) in value.items():
            ws.cell(row=row, column=column, value=value)
            fill = PatternFill(start_color=colors[section], end_color=colors[section], fill_type="solid")
            ws.cell(row=row, column=column).fill = fill
            row += 1
        column += 1
    
    #Fill in the length counters
    column = 20
    for (event,value) in length_counters.items():
        row = length_counter_row_start
        for pattern in value.values():
            ws.cell(row=row, column=column, value=pattern)
            row += 1
        column += 1
    
    column = 20
    high_amplitude_start = row
    for (event,value) in high_amplitude_counters.items():
        row = high_amplitude_start
        for pattern in value.values():
            ws.cell(row=row, column=column, value=pattern)
            row += 1
        column += 1

    print(counters)
    # Save the workbook
    wb.save(file_name)
    return wb